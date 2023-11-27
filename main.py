
from pathlib import Path
import ifcopenshell
import ifcopenshell.util.element
import ifcopenshell.api
import openpyxl
from openpyxl import load_workbook

modelname = "G16_Skylab_Model"

try:
    dir_path = Path(__file__).parent
    model_url = Path.joinpath(dir_path, 'model', modelname).with_suffix('.ifc')
    model = ifcopenshell.open(model_url)
except OSError:
    try:
        import bpy
        model_url = Path.joinpath(Path(bpy.context.space_data.text.filepath).parent, 'model', modelname).with_suffix('.ifc')
        model = ifcopenshell.open(model_url)
    except OSError:
        print(f"ERROR: please check your model folder : {model_url} does not exist")

# Open the Excel file
wb = openpyxl.Workbook()

#Load excel with densities:
excel_densities = load_workbook("List_of_densities.xlsx")
sheet_densities = excel_densities['Sheet1']
ecoinvent = excel_densities['Ecoinvent']
temp = sheet_densities["A"]
density_names = [temp[x].value for x in range(len(temp))]
temp = sheet_densities["B"]
density_values = [temp[x].value for x in range(len(temp))]
temp = sheet_densities["C"]
material_number_code = [temp[x].value for x in range(len(temp))]
temp = sheet_densities["D"]
ecoinvent_list_unit = [temp[x].value for x in range(len(temp))]
temp = ecoinvent["B"]
ecoinvent_numbers = [temp[x].value for x in range(len(temp))]
temp = ecoinvent["N"]
ecoinvent_IPCC2021 = [temp[x].value for x in range(len(temp))]

# Create two worksheets
ws_slabs = wb.create_sheet('Slabs')
ws_walls = wb.create_sheet('Walls')

#Calculate slab volumes/names
slab_material_names =[]
slab_material_volumes = []
for Slab in model.by_type("IfcSlab"):
    material = ifcopenshell.util.element.get_material(Slab)
    name = ifcopenshell.util.element.get_type(Slab)
    psets = ifcopenshell.util.element.get_pset(Slab,'Qto_SlabBaseQuantities')
    
    if material is not None:
        if material.Name not in slab_material_names:
            slab_material_names.append(material.Name)
            slab_material_volumes.append(float(psets['GrossVolume']))
            ws_slabs.cell(row=1+len(slab_material_names), column=3).value = '[m3]'
            ws_slabs.cell(row=1+len(slab_material_names), column=5).value = '[kg/m3]'
            ws_slabs.cell(row=1+len(slab_material_names), column=7).value = '[kg]'
            #Make it add the density the first time a material is encountered
        else:
            index = slab_material_names.index(material.Name)
            slab_material_volumes[index] = slab_material_volumes[index] + float(psets['GrossVolume'])
    else:
        print("Slab has no Name")

#Slab densities
slab_material_densities = [0] * len(slab_material_names)
for i in range(len(slab_material_names)):
    try:
        index = density_names.index(slab_material_names[i])
        slab_material_densities[i] = float(density_values[index])
    except TypeError:
        slab_material_densities[i] = "No density noted"    
    except ValueError:
        slab_material_densities[i] = "Material name not found in densities, or letters found in density" 

#Slab weight
slab_material_weight = []
for i in range(len(slab_material_names)):
    try:
        slab_material_weight.append(slab_material_densities[i]*slab_material_volumes[i])
    except TypeError:
        slab_material_weight.append("Weight cannot be calculated")

#Total slab emissions are calculated
slab_emission = [0] * len(slab_material_names)
slab_total_emissions = []
slab_emission_unit = [0] * len(slab_material_names)
for i in range(len(slab_material_names)):
    try:
        index = density_names.index(slab_material_names[i])
        slab_emission[i] = (material_number_code[index])
        slab_emission_unit[i] = (ecoinvent_list_unit[index])
    except TypeError:
        slab_emission[i] = "No ecoinvent material chosen"

for i in range(len(slab_material_names)):
    try:
        if 'm3' in slab_emission_unit[i]:
            index = ecoinvent_numbers.index(slab_emission[i])
            slab_total_emissions.append(float(ecoinvent_IPCC2021[index]) * float(slab_material_volumes[i]))
        if "kg" in slab_emission_unit[i]:
            index = ecoinvent_numbers.index(slab_emission[i])
            slab_total_emissions.append(float(ecoinvent_IPCC2021[index]) * float(slab_material_weight[i]))
    except ValueError:
        slab_total_emissions.append("Number not in list")
    except TypeError:
        slab_total_emissions.append("Something has gone horribly wrong")


for Slab in model.by_type("IfcSlab"):
    index = "Error"
    index2 = "Error"
    index3 = "Error"
    index4 = "Error"
    weight = "Error"

    material = ifcopenshell.util.element.get_material(Slab)
    psets = ifcopenshell.util.element.get_pset(Slab,'Qto_SlabBaseQuantities')
    try:
        index = slab_material_names.index(material[0])
    except TypeError:
        print("TypeError")
    try: 
        index2 = density_names.index(material[0])
        index5 = material_number_code[index2]
        index4 = ecoinvent_list_unit[index2]
        index2 = ecoinvent_numbers.index(index5)
        index3 = ecoinvent_IPCC2021[index2]
    except TypeError: 
        print("TypeError")
    try:
        weight = slab_material_densities[index] * psets['GrossVolume']
    except KeyError: 
        print("KeyError")
    except TypeError:
        print("TypeError")
    if 'm3' == index4:
        try:
            index2 = round(index3 * psets['GrossVolume'],2)
        except KeyError: 
            print("KeyError")
    if 'kg' == index4:
        try:
            index2 = round(index3 * weight,2)
        except KeyError: 
            print("KeyError")
    pset = ifcopenshell.api.run("pset.add_pset", model, product=Slab, name="Environdex")
    try:
        ifcopenshell.api.run("pset.edit_pset", model, pset=pset, properties={"Material": material[0] , "Volume": psets['GrossVolume'], "Density": slab_material_densities[index], "Weight": weight,"Ecoinvent process" : index5, "IPCC2021 Global Warming": index2})
    except TypeError:
        print("TypeError")

# Calculate wall volumes/names
wall_material_names =[]
wall_material_volumes = []

for Wall in model.by_type("IfcWall"):
    material = ifcopenshell.util.element.get_material(Wall)
    name = ifcopenshell.util.element.get_type(Wall)
    psets = ifcopenshell.util.element.get_pset(Wall,'Qto_WallBaseQuantities')
    
    if material is not None:
        if material.Name not in wall_material_names:
            wall_material_names.append(material.Name)
            wall_material_volumes.append(float(psets['GrossVolume']))
            ws_walls.cell(row=1+len(wall_material_names), column=3).value = '[m3]'
            ws_walls.cell(row=1+len(wall_material_names), column=5).value = '[kg/m3]'
            ws_walls.cell(row=1+len(wall_material_names), column=7).value = '[kg]'
            #Add density first time a material is found
        else:
            index = wall_material_names.index(material.Name)
            wall_material_volumes[index] = wall_material_volumes[index] + float(psets['GrossVolume'])
    else:
        print("Wall has no Name")

#List of wall densities is made
wall_material_densities = [0] *len(wall_material_names)
for i in range(len(wall_material_names)):
    try:
        index = density_names.index(wall_material_names[i])
        wall_material_densities[i] = float(density_values[index])
    except TypeError:
        wall_material_densities[i] = "No density noted" 
    except ValueError:
        wall_material_densities[i] = "Material name not found in densities, or letters found in density" 

#Wall weight
wall_material_weight = []
for i in range(len(wall_material_names)):
    try:
        wall_material_weight.append(wall_material_densities[i]*wall_material_volumes[i])
    except TypeError:
        wall_material_weight.append("Weight cannot be calculated")

#Total wall emissions are calculated
wall_emission = [0] * len(wall_material_names)
wall_total_emissions = []
wall_emission_unit = [0] * len(wall_material_names)
for i in range(len(wall_material_names)):
    try:
        index = density_names.index(wall_material_names[i])
        wall_emission[i] = (material_number_code[index])
        wall_emission_unit[i] = (ecoinvent_list_unit[index])
    except TypeError:
        wall_emission[i] = "No ecoinvent material chosen"

for i in range(len(wall_material_names)):
    try:
        if 'm3' in wall_emission_unit[i]:
            index = ecoinvent_numbers.index(wall_emission[i])
            wall_total_emissions.append(float(ecoinvent_IPCC2021[index]) * float(wall_material_volumes[i]))
        if "kg" in wall_emission_unit[i]:
            index = ecoinvent_numbers.index(wall_emission[i])
            wall_total_emissions.append(float(ecoinvent_IPCC2021[index]) * float(wall_material_weight[i]))
    except ValueError:
        wall_total_emissions.append("Number not in list")
    except TypeError:
        wall_total_emissions.append("Something has gone horribly wrong")

wall_CO2_unit = ["kg CO2-eq"]*len(wall_material_names)
slab_CO2_unit = ["kg CO2-eq"]*len(slab_material_names)

for wall in model.by_type("IfcWall"):
    index = "Error"
    index2 = "Error"
    index3 = "Error"
    index4 = "Error"
    weight = "Error"

    material = ifcopenshell.util.element.get_material(wall)
    psets = ifcopenshell.util.element.get_pset(wall,'Qto_WallBaseQuantities')
    try:
        index = wall_material_names.index(material[0])
    except TypeError:
        print("TypeError")
    try: 
        index2 = density_names.index(material[0])
        index5 = material_number_code[index2]
        index4 = ecoinvent_list_unit[index2]
        index2 = ecoinvent_numbers.index(index5)
        index3 = ecoinvent_IPCC2021[index2]
    except TypeError: 
        print("TypeError")
    try:
        weight = wall_material_densities[index] * psets['GrossVolume']
    except KeyError: 
        print("KeyError")
    except TypeError:
        print("TypeError")
    if 'm3' == index4:
        try:
            index2 = round(index3 * psets['GrossVolume'],2)
        except KeyError: 
            print("KeyError")
        except TypeError:
            print("TypeError")
            
    if 'kg' == index4:
        try:
            index2 = round(index3 * weight,2)
        except KeyError: 
            print("KeyError")
        except TypeError:
            print(TypeError)
    pset = ifcopenshell.api.run("pset.add_pset", model, product=wall, name="Environdex")
    try:
        ifcopenshell.api.run("pset.edit_pset", model, pset=pset, properties={"Material": material[0] , "Volume": psets['GrossVolume'], "Density": wall_material_densities[index], "Weight": weight,"Ecoinvent process" : index5, "IPCC2021 Global Warming": index2})
    except TypeError:
        print("TypeError")

# Write the header row to the worksheets
ws_slabs.cell(row=1, column=1).value = 'Material Name'
ws_slabs.cell(row=1, column=2).value = 'Volume'
ws_slabs.cell(row=1, column=3).value = 'Unit'
ws_slabs.cell(row=1, column=4).value = 'Density'
ws_slabs.cell(row=1, column=5).value = 'Unit'
ws_slabs.cell(row=1, column=6).value = 'Weight'
ws_slabs.cell(row=1, column=7).value = 'Unit'
ws_slabs.cell(row=1, column=8).value = 'IPCC2021 Climate change'
ws_slabs.cell(row=1, column=9).value = 'Unit'
ws_walls.cell(row=1, column=1).value = 'Material Name'
ws_walls.cell(row=1, column=2).value = 'Volume'
ws_walls.cell(row=1, column=3).value = 'Unit'
ws_walls.cell(row=1, column=4).value = 'Densities'
ws_walls.cell(row=1, column=5).value = 'Unit'
ws_walls.cell(row=1, column=6).value = 'Weight'
ws_walls.cell(row=1, column=7).value = 'Unit'
ws_walls.cell(row=1, column=8).value = 'IPCC2021 Climate change'
ws_walls.cell(row=1, column=9).value = 'Unit'

# Write the data to the worksheets
row = 2
for slab_material_names, slab_material_volumes, slab_material_densities, slab_material_weight, slab_total_emissions, slab_CO2_unit in zip(slab_material_names, slab_material_volumes, slab_material_densities, slab_material_weight, slab_total_emissions,slab_CO2_unit):
    ws_slabs.cell(row=row, column=1).value = slab_material_names #Names
    ws_slabs.cell(row=row, column=2).value = slab_material_volumes #Volumes
    ws_slabs.cell(row=row, column=4).value = slab_material_densities #Densities
    ws_slabs.cell(row=row, column=6).value = slab_material_weight #Weight
    ws_slabs.cell(row=row, column=8).value = slab_total_emissions 
    ws_slabs.cell(row=row, column=9).value = slab_CO2_unit
    row += 1

row = 2
for wall_material_names, wall_material_volumes, wall_material_densities, wall_material_weight, wall_total_emissions, wall_CO2_unit in zip(wall_material_names, wall_material_volumes, wall_material_densities, wall_material_weight, wall_total_emissions, wall_CO2_unit):
    ws_walls.cell(row=row, column=1).value = wall_material_names #Names
    ws_walls.cell(row=row, column=2).value = wall_material_volumes #Volumes
    ws_walls.cell(row=row, column=4).value = wall_material_densities #Densities
    ws_walls.cell(row=row, column=6).value = wall_material_weight #Weight
    ws_walls.cell(row=row, column=8).value = wall_total_emissions #Weight
    ws_walls.cell(row=row, column=9).value = wall_CO2_unit
    row += 1

# Save the Excel file
wb.save('material_quantities.xlsx')
model.write(r'C:\Users\madsf\OneDrive\Skrivebord\Advanced BIM\Model/G16_Skylab_Modified.ifc')

print("Operation Complete")
